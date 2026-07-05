"""
模块4 · 全链路输出：分类 → 提取 → Excel 生成

用法：
    python3 phase1/generate_report.py
    
输出：
    output/差旅报销汇总_YYYY-MM-DD.xlsx
    output/附件/
"""

import imaplib, email, os, re, json
from email.header import decode_header
from datetime import datetime
from dotenv import load_dotenv
from io import BytesIO

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

# ========== 配置 ==========
PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

OUTPUT_DIR = os.path.join(PROJECT_DIR, 'output')
try:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(os.path.join(OUTPUT_DIR, '附件'), exist_ok=True)
    test_file = os.path.join(OUTPUT_DIR, '.write_test')
    with open(test_file, 'w') as f: f.write('ok')
    os.remove(test_file)
except PermissionError:
    print(f"❌ 沙箱限制：无法写入项目目录 {OUTPUT_DIR}")
    print(f"📌 请手动运行：cp -r /tmp/biztrip-agent_output/* \"{OUTPUT_DIR}/\"")
    exit(1)

# ========== 复用模块2的分类规则 ==========
DOMAIN_RULES = {
    '机票': ['qunar.com', 'ctrip.com', 'fliggy.com', 'alitrip.com', 'airchina', 'ceair.com', 'csair.com', 'trip.com'],
    '火车票': ['12306.cn', 'zhixing.com', 'tiexing.com'],
    '酒店': ['booking.com', 'agoda.com', 'airbnb.com', 'meituan.com', 'huazhuhotels.com'],
    '网约车': ['didiglobal.com', 'xiaojukeji.com', 'uber.com', 'amap.com'],
    '发票': ['crestv.cn', 'fapiao.com', 'invoice.', 'txffp.com'],
    '门票': ['damai.cn', 'maoyan.com', 'showstart.com'],
}

KEYWORD_RULES = {
    '机票': ['机票', '航班', '登机', '航空'],
    '火车票': ['火车票', '高铁', '动车', '12306'],
    '酒店': ['酒店', '民宿', '入住'],
    '网约车': ['滴滴', '网约车'],
    '发票': ['发票', '报销', '电子凭证'],
}

SPAM_DOMAINS = ['job51', 'steampowered', 'email.apple.com', '10000@', 'cmbchina', '2ksports', 'amazon']


def decode_str(s):
    if not s: return ''
    parts = decode_header(s)
    result = []
    for content, charset in parts:
        if isinstance(content, bytes):
            try: result.append(content.decode(charset or 'utf-8', errors='replace'))
            except: result.append(content.decode('utf-8', errors='replace'))
        else: result.append(content)
    return ''.join(result)


def get_email_text(msg):
    """提取邮件正文 + PDF附件内容"""
    body = ''
    pdf_texts = []
    
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            raw = part.get_payload(decode=True)
            if not raw: continue
            
            if ct == 'text/plain':
                for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030']:
                    try: body = raw.decode(enc); break
                    except: pass
                if body: break
            
            if ct == 'text/html' and not body:
                for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030']:
                    try: html = raw.decode(enc); break
                    except: pass
                else: html = raw.decode('utf-8', errors='replace')
                html = re.sub(r'<(style|script)[^>]*>.*?</\1>', '', html, flags=re.DOTALL)
                html = re.sub(r'<[^>]+>', '\n', html)
                html = re.sub(r'\n{3,}', '\n\n', html)
                html = re.sub(r'&nbsp;', ' ', html)
                body = html.strip()
    
    # PDF附件
    if msg.is_multipart():
        for part in msg.walk():
            fn = part.get_filename()
            ct = part.get_content_type()
            fn_decoded = decode_str(fn) if fn else ''
            is_pdf = (fn and fn.lower().endswith('.pdf')) or ct == 'application/pdf' or '.pdf' in fn_decoded.lower()
            if is_pdf:
                raw = part.get_payload(decode=True)
                if raw:
                    try:
                        from PyPDF2 import PdfReader
                        reader = PdfReader(BytesIO(raw))
                        for page in reader.pages:
                            t = page.extract_text()
                            if t: pdf_texts.append(t)
                    except:
                        pass
    
    # ZIP附件 → 解压后找内部PDF提取文本
    if msg.is_multipart():
        for part in msg.walk():
            fn = part.get_filename()
            if not fn: continue
            fn_decoded = decode_str(fn)
            if not fn_decoded.lower().endswith('.zip'): continue
            raw = part.get_payload(decode=True)
            if not raw: continue
            try:
                import zipfile
                zf = zipfile.ZipFile(BytesIO(raw))
                for info in zf.infolist():
                    if not info.filename.lower().endswith('.pdf'): continue
                    pdf_data = zf.read(info.filename)
                    from PyPDF2 import PdfReader
                    reader = PdfReader(BytesIO(pdf_data))
                    for page in reader.pages:
                        t = page.extract_text()
                        if t: pdf_texts.append(t)
                zf.close()
            except:
                pass
    
    if pdf_texts:
        body = body + '\n' + '\n'.join(pdf_texts)
    return body or ''


def save_attachment(msg, email_idx):
    """保存符合类型的附件（PDF + 非网页资源图片）到 output/附件/"""
    KEEP_EXTS = ('.pdf', '.zip', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.heic')
    IMG_BLOCKLIST = (
        'advertising', 'downloadbutton', 'header', 'footer', 'logo', 'banner',
        'icon', 'spacer', 'cta', 'button', 'social', 'wechat', 'weibo', 'qrcode',
    )
    saved = []
    if msg.is_multipart():
        for part in msg.walk():
            fn = part.get_filename()
            if not fn: continue
            fn_decoded = decode_str(fn)
            fn_lower = fn_decoded.lower()
            if not fn_lower.endswith(KEEP_EXTS):
                continue
            # 图片文件过滤：排除网页广告/按钮/头图等
            if fn_lower.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.heic')):
                base_name = os.path.splitext(fn_lower)[0]
                if any(bl in base_name for bl in IMG_BLOCKLIST):
                    continue
            raw = part.get_payload(decode=True)
            if raw:
                safe_name = f"{email_idx}_{fn_decoded}"
                path = os.path.join(OUTPUT_DIR, '附件', safe_name)
                with open(path, 'wb') as f:
                    f.write(raw)
                saved.append(safe_name)
    return saved


def classify(sender, subject):
    """基于规则分类（同模块2）"""
    sl = sender.lower()
    if any(d in sl for d in SPAM_DOMAINS):
        return None
    for cat, domains in DOMAIN_RULES.items():
        for d in domains:
            if d in sl: return cat
    text = subject.lower()
    for cat, keywords in KEYWORD_RULES.items():
        for kw in keywords:
            if kw in text: return cat
    return None


def get_pdf_attachments(msg):
    """提取邮件中所有 PDF 附件的（原始文件名，文本内容）"""
    pdfs = []
    if not msg.is_multipart(): return pdfs
    for part in msg.walk():
        fn = part.get_filename()
        if not fn: continue
        fn_decoded = decode_str(fn)
        ct = part.get_content_type()
        is_pdf = (fn.lower().endswith('.pdf')) or ct == 'application/pdf' or '.pdf' in fn_decoded.lower()
        if not is_pdf: continue
        raw = part.get_payload(decode=True)
        if not raw: continue
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(BytesIO(raw))
            text = '\n'.join(page.extract_text() for page in reader.pages if page.extract_text())
        except:
            text = ''
        pdfs.append({'filename': fn_decoded, 'text': text})
    return pdfs


def extract_pdf_info(pdf):
    """从 PDF 文件名和文本提取机票/酒店/网约车结构化信息"""
    info = {'日期': '', '出发地': '', '目的地': '', '金额': ''}
    fn = pdf['filename']
    text = pdf['text']

    # 1) 从文件名解析日期：2026-05-30 / 2026年05月30日
    m = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', fn)
    if m:
        info['日期'] = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"

    # 2) 从文件名解析路线：成都-甘孜
    m = re.search(r'([\u4e00-\u9fa5]{2,4})-([\u4e00-\u9fa5]{2,4})', fn)
    if m:
        info['出发地'] = m.group(1)
        info['目的地'] = m.group(2)

    # 3) 从文件名解析金额：文件名末尾 1755.00.pdf
    m = re.search(r'(\d+\.\d{2})\.pdf$', fn)
    if m:
        info['金额'] = float(m.group(1))

    # 4) 从 PDF 文本补充金额（优先价税合计，其次 ¥xxx）
    if not info['金额']:
        m = re.search(r'价税合计[（(]小写[）)]\s*[¥￥](\d+\.?\d*)', text)
        if m:
            info['金额'] = float(m.group(1))
        else:
            m = re.search(r'[¥￥]\s*(\d+\.?\d*)', text)
            if m:
                info['金额'] = float(m.group(1))

    # 5) 从 PDF 文本补充日期
    if not info['日期']:
        m = re.search(r'(\d{4}年\d{1,2}月\d{1,2}日)', text)
        if m:
            info['日期'] = m.group(1)

    return info


def extract_fields(body, subject, sender, category, attachments, pdf_info=None):
    """基于规则提取（同模块3），支持用 PDF 信息覆盖"""
    data = {
        '分类': category, '平台': '', '金额': '', '日期': '',
        '供应商': '', '出发地': '', '目的地': '',
        '附件': '; '.join(attachments) if attachments else ''
    }
    
    # 识别平台
    platform_map = {
        'qunar.com': '去哪儿网', 'ctrip.com': '携程', 'xiaojukeji.com': '滴滴',
        'crestv.cn': '智慧发票', 'huazhuhotels.com': '华住酒店', 'amap.com': '高德',
        'txffp.com': '票根', '12306': '12306', 'fliggy.com': '飞猪',
    }
    for domain, name in platform_map.items():
        if domain in sender.lower():
            data['平台'] = name
            break
    
    # PDF 信息覆盖（若从 PDF 文件名/文本已解析出准确值）
    if pdf_info:
        if pdf_info.get('金额'): data['金额'] = pdf_info['金额']
        if pdf_info.get('日期'): data['日期'] = pdf_info['日期']
        if pdf_info.get('出发地'): data['出发地'] = pdf_info['出发地']
        if pdf_info.get('目的地'): data['目的地'] = pdf_info['目的地']
    
    # 金额 —— 若 PDF 没提供，再按模式匹配
    if not data['金额']:
        patterns = [
            r'合[计]\s*(\d+\.?\d*)\s*元',
            r'发票金额[：:]\s*(\d+\.?\d*)',
            r'金额[：:]\s*(\d+\.?\d*)',
            r'价[格款][：:]\s*(\d+\.?\d*)',
            r'价税合计[（(]小写[）)]\s*[¥￥]?\s*(\d+\.?\d*)',
            r'合[计]\s*[¥￥]?\s*(\d+\.?\d*)',
            r'(\d+\.\d{2})\s*元',
            r'(\d+\.?\d*)\s*元',
            r'[¥￥]\s*(\d+\.?\d*)',
        ]
        for p in patterns:
            m = re.search(p, body)
            if m: data['金额'] = float(m.group(1)); break
        if not data['金额']:
            m = re.search(r'金额[：:](\d+\.?\d*)', subject)
            if m: data['金额'] = float(m.group(1))
    
    # 金额终极兜底：从PDF文本中找所有 XX.XX 数字，取最大值（用于12306等编码异常的PDF）
    if not data['金额']:
        all_amounts = [float(m.group(1)) for m in re.finditer(r'(\d+\.\d{2})', body) if float(m.group(1)) > 1]
        if all_amounts:
            data['金额'] = max(all_amounts)
    
    # 日期 —— 若 PDF 没提供，再按模式匹配
    if not data['日期']:
        for p in [r'(\d{4}年\d{1,2}月\d{1,2}日)', r'(\d{4}[-/年]\d{1,2}[-/月]\d{1,2})']:
            m = re.search(p, body)
            if m: data['日期'] = m.group(1); break
    
    # 出发地→目的地 —— 若 PDF 没提供，再按模式匹配
    if not data['出发地'] or not data['目的地']:
        m = re.search(r'([\u4e00-\u9fa5]{2,4})\s*[→\-]\s*([\u4e00-\u9fa5]{2,4})', body)
        if m: data['出发地'], data['目的地'] = m.group(1), m.group(2)
    
    return data


def get_date_filter():
    """交互：询问用户日期范围，返回 (search_cmd, display_str) 或 None"""
    print("\n📅 是否指定搜索日期范围？（可选，直接回车按默认搜索）")
    start = input("  开始日期 (如 2026-05-27): ").strip()
    end = input("  结束日期 (如 2026-06-05): ").strip()
    
    if not start and not end:
        return None
    
    parts = []
    display = []
    if start:
        try:
            dt = datetime.strptime(start, '%Y-%m-%d')
            parts.append(f'SINCE {dt.strftime("%d-%b-%Y")}')
            display.append(f'{start}')
        except:
            print(f"  ⚠️ 开始日期格式错误（需要 YYYY-MM-DD），忽略")
    if end:
        try:
            dt = datetime.strptime(end, '%Y-%m-%d')
            parts.append(f'BEFORE {dt.strftime("%d-%b-%Y")}')
            display.append(f'{end}')
        except:
            print(f"  ⚠️ 结束日期格式错误（需要 YYYY-MM-DD），忽略")
    
    if parts:
        return (' '.join(parts), '~'.join(display))
    return None


def _get_email_config():
    """获取邮箱配置，支持通用变量 + 向后兼容旧 QQ_EMAIL"""
    account = os.getenv('EMAIL_ACCOUNT', '') or os.getenv('QQ_EMAIL', '')
    password = os.getenv('EMAIL_PASSWORD', '') or os.getenv('QQ_AUTH_CODE', '')
    server = os.getenv('EMAIL_IMAP_SERVER', '') or os.getenv('QQ_IMAP_SERVER', '')
    port = int(os.getenv('EMAIL_IMAP_PORT', '') or os.getenv('QQ_IMAP_PORT', '0'))
    if not server:
        if '@163.com' in account: server = 'imap.163.com'
        elif '@126.com' in account: server = 'imap.126.com'
        elif '@gmail.com' in account: server = 'imap.gmail.com'
        elif '@outlook.com' in account or '@hotmail.com' in account: server = 'outlook.office365.com'
        else: server = 'imap.qq.com'
    if not port: port = 993
    return account, password, server, port


# ========== 主流程 ==========
def main():
    email_addr, auth_code, imap_server, imap_port = _get_email_config()
    
    if not email_addr or not auth_code:
        print("❌ 请在 .env 中配置邮箱信息")
        print("   EMAIL_ACCOUNT=your_email@example.com")
        print("   EMAIL_PASSWORD=your_authorization_code")
        return
    
    try:
        conn = imaplib.IMAP4_SSL(imap_server, imap_port)
        conn.login(email_addr, auth_code)
        conn.select('INBOX')
        
        # ---- 询问日期范围 ----
        date_filter = get_date_filter()
        
        if date_filter:
            search_cmd, date_label = date_filter
            print(f"📅 按日期搜索: {date_label}")
            status, data = conn.search(None, search_cmd)
            if status != 'OK' or not data[0]:
                print("   无匹配邮件，结束")
                return
            mail_ids = data[0].split()
            print(f"   匹配 {len(mail_ids)} 封邮件")
            scan_label = f'日期范围 {date_label}'
        else:
            print("📅 未指定日期范围 → 默认按最近 60 封搜索")
            status, data = conn.search(None, 'ALL')
            if status != 'OK' or not data[0]:
                print("收件箱为空")
                return
            mail_ids = data[0].split()
            mail_ids = mail_ids[-60:]
            scan_label = '最近60封'
        
        records = []
        
        print(f"正在扫描 {len(mail_ids)} 封邮件...")
        
        for idx, mid in enumerate(reversed(mail_ids), 1):
            status, msg_data = conn.fetch(mid, '(RFC822)')
            if status != 'OK': continue
            
            msg = email.message_from_bytes(msg_data[0][1])
            sender = decode_str(msg['From'])
            subject = decode_str(msg['Subject'])
            
            cat = classify(sender, subject)
            if not cat:
                continue
            
            body = get_email_text(msg)
            attachments = save_attachment(msg, idx)
            
            # 机票类邮件：按 PDF 附件拆分为多条记录，并从文件名提取准确信息
            if cat == '机票':
                pdfs = get_pdf_attachments(msg)
                if pdfs:
                    for pdf in pdfs:
                        pdf_info = extract_pdf_info(pdf)
                        record = extract_fields(body, subject, sender, cat, attachments, pdf_info=pdf_info)
                        records.append(record)
                        print(f"  [{idx}] {cat}: {subject[:30]} | {pdf['filename'][:35]} | {record['金额']}")
                    continue
            
            record = extract_fields(body, subject, sender, cat, attachments)
            records.append(record)
            
            print(f"  [{idx}] {cat}: {subject[:40]} {record['金额']}")
        
        conn.logout()
        
        if not records:
            print("未发现出差相关邮件")
            return
        
        # ---------- 生成 Excel ----------
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # 颜色主题
        BLUE = '1A56DB'
        LIGHT_BLUE = 'E8EFFB'
        WHITE = 'FFFFFF'
        LIGHT_GRAY = 'F3F4F6'
        DARK_TEXT = '1F2937'
        
        header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
        header_font = Font(bold=True, color=WHITE, size=11, name='微软雅黑')
        title_font = Font(bold=True, color=DARK_TEXT, size=14, name='微软雅黑')
        body_font = Font(size=10, name='微软雅黑')
        bold_font = Font(bold=True, size=10, name='微软雅黑')
        total_font = Font(bold=True, color=BLUE, size=11, name='微软雅黑')
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        
        # 分类颜色
        cat_fills = {
            '机票': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
            '酒店': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
            '网约车': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
            '发票': PatternFill(start_color='FCE7F3', end_color='FCE7F3', fill_type='solid'),
            '火车票': PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid'),
        }
        
        total = sum(r['金额'] for r in records if r['金额'] != '')
        
        # ===== Sheet 1: 报销总览 =====
        ws = wb.active
        ws.title = '报销总览'
        ws.sheet_properties.tabColor = BLUE
        
        # 标题
        ws.merge_cells('A1:F1')
        ws['A1'] = '差旅费用报销汇总单'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 报告日期
        ws.merge_cells('A2:F2')
        ws['A2'] = f'报告生成日期：{datetime.now().strftime("%Y年%m月%d日")}  |  扫描范围：{scan_label}  |  出差相关：{len(records)}封'
        ws['A2'].font = Font(size=9, color='6B7280', name='微软雅黑')
        ws.row_dimensions[2].height = 20
        
        # 总计卡片
        ws.merge_cells('A4:C4')
        ws['A4'] = '💰 可报销总额'
        ws['A4'].font = Font(size=12, color='6B7280', name='微软雅黑')
        ws.merge_cells('D4:F4')
        ws['D4'] = f'¥ {total:,.2f}'
        ws['D4'].font = Font(bold=True, size=22, color=BLUE, name='微软雅黑')
        ws['D4'].alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[4].height = 40
        
        # 统计指标
        cats_with_amount = len([r for r in records if r['金额'] != ''])
        avg_per_item = total / cats_with_amount if cats_with_amount > 0 else 0
        stats = [
            ('费用笔数', f'{cats_with_amount} 笔', '分类数', f'{len(set(r["分类"] for r in records))} 类'),
            ('单笔均额', f'¥ {avg_per_item:.2f}', '无金额记录', f'{len(records) - cats_with_amount} 笔'),
        ]
        for i, (label1, val1, label2, val2) in enumerate(stats):
            row = 5 + i
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = label1
            ws[f'A{row}'].font = Font(size=10, color='6B7280', name='微软雅黑')
            ws[f'D{row}'] = val1
            ws[f'D{row}'].font = Font(bold=True, size=11, color=DARK_TEXT, name='微软雅黑')
            ws.merge_cells(f'E{row}:F{row}')
            ws[f'E{row}'] = label2
            ws[f'E{row}'].font = Font(size=10, color='6B7280', name='微软雅黑')
        
        # 分类汇总表
        row = 8
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = '📊 按类别汇总'
        ws[f'A{row}'].font = Font(bold=True, size=11, color=DARK_TEXT, name='微软雅黑')
        row = 9
        
        cat_headers = ['类别', '笔数', '金额(元)', '占比', '', '']
        cat_widths = [12, 10, 14, 10, 10, 10]
        for col, (h, w) in enumerate(zip(cat_headers, cat_widths), 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w
        
        categories = {}
        for r in records:
            if r['金额'] != '':
                cat = r['分类']
                categories[cat] = categories.get(cat, 0) + r['金额']
        
        for cat, amt in sorted(categories.items(), key=lambda x: -x[1]):
            row += 1
            count = sum(1 for r in records if r['分类'] == cat and r['金额'] != '')
            vals = [cat, count, amt, f'{amt/total*100:.1f}%', '', '']
            fill = cat_fills.get(cat, PatternFill())
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = body_font
                cell.alignment = center_align
                cell.border = thin_border
                if cat in cat_fills:
                    cell.fill = cat_fills[cat]
        
        # 合计行
        row += 1
        total_count = sum(1 for r in records if r['金额'] != '')
        vals = ['合计', total_count, total, '100%', '', '']
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = total_font
            cell.fill = PatternFill(start_color='E8EFFB', end_color='E8EFFB', fill_type='solid')
            cell.alignment = center_align
            cell.border = thin_border
        
        # ===== Sheet 2: 费用明细 =====
        ws2 = wb.create_sheet('费用明细')
        ws2.sheet_properties.tabColor = '3B82F6'
        
        # 标题
        ws2.merge_cells('A1:G1')
        ws2['A1'] = '差旅费用明细表'
        ws2['A1'].font = title_font
        ws2.row_dimensions[1].height = 28
        
        # 表头
        headers = ['序号', '日期', '类别', '供应商/平台', '金额(元)', '出发地→目的地', '附件文件']
        widths = [6, 14, 10, 18, 12, 18, 20]
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws2.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws2.column_dimensions[get_column_letter(col)].width = w
        
        # 数据行（按日期排序）
        sorted_records = sorted(records, key=lambda r: str(r.get('日期', '')), reverse=True)
        
        for i, r in enumerate(sorted_records, 1):
            row = i + 2
            route = f"{r.get('出发地', '')}→{r.get('目的地', '')}" if r.get('出发地') and r.get('目的地') else ''
            
            vals = [
                i,
                r['日期'] or '-',
                r['分类'],
                r['平台'] or '-',
                r['金额'] if r['金额'] != '' else '-',
                route or '-',
                r['附件'] or '-'
            ]
            
            cat_fill = cat_fills.get(r['分类'], PatternFill())
            row_fill = cat_fill if r['分类'] in cat_fills else PatternFill()
            
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.font = body_font
                cell.alignment = center_align if col in (1, 3, 5) else left_align
                cell.border = thin_border
                if col == 3:
                    cell.fill = cat_fill
        
        # 最高金额高亮
        max_amt = max((r['金额'] for r in sorted_records if r['金额'] != ''), default=0)
        for r_idx, r in enumerate(sorted_records, 3):
            if r['金额'] == max_amt:
                for col in range(1, 8):
                    ws2.cell(row=r_idx, column=col).font = Font(bold=True, size=10, name='微软雅黑', color='DC2626')
        
        # ===== Sheet 3: 汇总统计 =====
        ws3 = wb.create_sheet('按供应商')
        ws3.sheet_properties.tabColor = '10B981'
        
        ws3.merge_cells('A1:D1')
        ws3['A1'] = '按供应商/平台统计'
        ws3['A1'].font = title_font
        ws3.row_dimensions[1].height = 28
        
        vendor_headers = ['供应商', '笔数', '金额(元)', '占比']
        vendor_widths = [20, 10, 14, 10]
        for col, (h, w) in enumerate(zip(vendor_headers, vendor_widths), 1):
            cell = ws3.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws3.column_dimensions[get_column_letter(col)].width = w
        
        # 按平台汇总
        vendors = {}
        for r in records:
            v = r['平台'] or '其他'
            if r['金额'] != '':
                if v not in vendors:
                    vendors[v] = {'count': 0, 'amount': 0}
                vendors[v]['count'] += 1
                vendors[v]['amount'] += r['金额']
        
        for i, (v, data) in enumerate(sorted(vendors.items(), key=lambda x: -x[1]['amount']), 1):
            row = i + 2
            vals = [v, data['count'], data['amount'], f"{data['amount']/total*100:.1f}%"]
            row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid') if i % 2 == 0 else PatternFill()
            for col, val in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=col, value=val)
                cell.font = body_font
                cell.alignment = center_align
                cell.border = thin_border
                cell.fill = row_fill
        
        # 保存
        today = datetime.now().strftime('%Y%m%d')
        xlsx_path = os.path.join(OUTPUT_DIR, f'差旅汇总_{today}.xlsx')
        wb.save(xlsx_path)
        
        # 保存
        today = datetime.now().strftime('%Y%m%d')
        xlsx_path = os.path.join(OUTPUT_DIR, f'差旅汇总_{today}.xlsx')
        wb.save(xlsx_path)
        
        print(f"\n{'='*50}")
        print(f"✅ 报告已生成")
        print(f"📄 邮件数: {len(records)} 封")
        print(f"💰 可报销总额: {total:.2f} 元")
        print(f"📊 Excel: {xlsx_path}")
        print(f"📎 附件: output/附件/")
        
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
