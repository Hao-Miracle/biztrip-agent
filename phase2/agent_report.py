"""
BizTrip Agent · 主入口（Agent 模式）

自动检测 LLM 可用性，在「AI 增强模式」和「规则模式」之间自动切换。

用法：
    python3 phase2/agent_report.py

前置条件：
    - .env 配置 EMAIL_ACCOUNT + EMAIL_PASSWORD（必需）
    - .env 配置 EMAIL_IMAP_SERVER + EMAIL_IMAP_PORT（可选，有默认值）
    - .env 配置 LLM_API_KEY（可选，不配则纯规则模式）

输出：
    output/差旅汇总_YYYYMMDD.xlsx  （三 Sheet Excel + 出差聚合）
    output/附件/                    （原始 PDF/ZIP 附件）

模式说明：
    🧠 AI 增强模式：LLM 智能分类 + LLM 专用提取 + LLM 出差聚合 + 规则降级兜底
    📋 规则模式：域名匹配 + 正则提取（零 API Key，和 Phase 1 相同）

支持模型：
    任何兼容 OpenAI 协议的服务商（DeepSeek / 通义千问 / GLM / Kimi / Ollama 本地模型等）
    在 .env 中配置 LLM_API_KEY + LLM_BASE_URL + LLM_MODEL 即可
"""

import imaplib
import email
import os
import sys
import re
import json
from email.header import decode_header
from datetime import datetime
from io import BytesIO

from dotenv import load_dotenv

# Phase 2 LLM 模块
from llm_classify import classify_email, has_api_key as llm_available
from llm_extract import extract_record
from llm_aggregate import aggregate_trips

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJECT_DIR, 'output')
ATTACH_DIR = os.path.join(OUTPUT_DIR, '附件')

os.makedirs(ATTACH_DIR, exist_ok=True)

# ========== 邮件解析工具 ==========


def decode_str(s):
    if not s:
        return ''
    parts = decode_header(s)
    result = []
    for content, charset in parts:
        if isinstance(content, bytes):
            try:
                result.append(content.decode(charset or 'utf-8', errors='replace'))
            except Exception:
                result.append(content.decode('utf-8', errors='replace'))
        else:
            result.append(content)
    return ''.join(result)


def get_email_text(msg):
    """提取邮件正文 + PDF/ZIP 附件文本"""
    body = ''
    pdf_texts = []

    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            raw = part.get_payload(decode=True)
            if not raw:
                continue

            if ct == 'text/plain':
                for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030']:
                    try:
                        body = raw.decode(enc)
                        break
                    except Exception:
                        pass
                if body:
                    break

            if ct == 'text/html' and not body:
                for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030']:
                    try:
                        html = raw.decode(enc)
                        break
                    except Exception:
                        pass
                else:
                    html = raw.decode('utf-8', errors='replace')
                html = re.sub(r'<(style|script)[^>]*>.*?</\1>', '', html, flags=re.DOTALL)
                html = re.sub(r'<[^>]+>', '\n', html)
                html = re.sub(r'\n{3,}', '\n\n', html)
                html = re.sub(r'&nbsp;', ' ', html)
                body = html.strip()

    # PDF / ZIP 附件
    if msg.is_multipart():
        for part in msg.walk():
            fn_raw = part.get_filename()
            fn = decode_str(fn_raw) if fn_raw else ''
            ct = part.get_content_type()
            raw = part.get_payload(decode=True)
            if not raw:
                continue

            if fn.lower().endswith('.pdf') or ct == 'application/pdf':
                try:
                    from PyPDF2 import PdfReader
                    reader = PdfReader(BytesIO(raw))
                    for page in reader.pages:
                        t = page.extract_text()
                        if t:
                            pdf_texts.append(t)
                except Exception:
                    pass

            if fn.lower().endswith('.zip'):
                try:
                    import zipfile
                    zf = zipfile.ZipFile(BytesIO(raw))
                    for info in zf.infolist():
                        if not info.filename.lower().endswith('.pdf'):
                            continue
                        pdf_data = zf.read(info.filename)
                        from PyPDF2 import PdfReader
                        reader = PdfReader(BytesIO(pdf_data))
                        for page in reader.pages:
                            t = page.extract_text()
                            if t:
                                pdf_texts.append(t)
                    zf.close()
                except Exception:
                    pass

    if pdf_texts:
        body = body + '\n' + '\n'.join(pdf_texts)
    return body or ''


def save_attachments(msg, email_idx):
    """保存附件"""
    KEEP_EXTS = ('.pdf', '.zip', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.heic')
    IMG_BLOCK = ('advertising', 'downloadbutton', 'header', 'footer', 'logo', 'banner',
                 'icon', 'spacer', 'cta', 'button', 'social', 'wechat', 'weibo', 'qrcode')
    saved = []

    if msg.is_multipart():
        for part in msg.walk():
            fn_raw = part.get_filename()
            if not fn_raw:
                continue
            fn = decode_str(fn_raw)
            fn_lower = fn.lower()
            if not fn_lower.endswith(KEEP_EXTS):
                continue
            if fn_lower.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.heic')):
                base_name = os.path.splitext(fn_lower)[0]
                if any(bl in base_name for bl in IMG_BLOCK):
                    continue
            raw = part.get_payload(decode=True)
            if raw:
                safe_name = f'{email_idx}_{fn}'
                path = os.path.join(ATTACH_DIR, safe_name)
                with open(path, 'wb') as f:
                    f.write(raw)
                saved.append(safe_name)
    return saved


# ========== 邮箱配置工具 ==========

def _get_email_config():
    """获取邮箱配置，支持通用变量 + 向后兼容旧 QQ_EMAIL 变量"""
    # 优先新变量，向后兼容旧 QQ_EMAIL
    account = os.getenv('EMAIL_ACCOUNT', '') or os.getenv('QQ_EMAIL', '')
    password = os.getenv('EMAIL_PASSWORD', '') or os.getenv('QQ_AUTH_CODE', '')

    # 根据账号自动推断 IMAP 服务器，也可通过 EMAIL_IMAP_SERVER 手动指定
    server = os.getenv('EMAIL_IMAP_SERVER', '') or os.getenv('QQ_IMAP_SERVER', '')
    port = int(os.getenv('EMAIL_IMAP_PORT', '') or os.getenv('QQ_IMAP_PORT', '0'))

    if not server:
        # 自动推断
        if '@qq.com' in account:
            server = 'imap.qq.com'
        elif '@163.com' in account:
            server = 'imap.163.com'
        elif '@126.com' in account:
            server = 'imap.126.com'
        elif '@gmail.com' in account:
            server = 'imap.gmail.com'
        elif '@outlook.com' in account or '@hotmail.com' in account:
            server = 'outlook.office365.com'
        else:
            server = 'imap.qq.com'  # 兜底

    if not port:
        port = 993

    return account, password, server, port


# ========== 主流程 ==========


def main():
    email_addr, auth_code, imap_server, imap_port = _get_email_config()

    if not email_addr or not auth_code:
        print('❌ 请在 .env 中配置邮箱信息')
        print('   EMAIL_ACCOUNT=your_email@example.com')
        print('   EMAIL_PASSWORD=your_authorization_code')
        return

    use_llm = llm_available()
    print('=' * 60)
    print('  BizTrip Agent')
    print(f'  模式: {"🧠 AI 增强模式" if use_llm else "📋 规则模式（零 API Key）"}')
    print(f'  邮箱: {email_addr} ({imap_server}:{imap_port})')
    if not use_llm:
        print('  💡 配置 LLM_API_KEY 即可启用 AI 增强（支持 DeepSeek/通义千问/GLM/Kimi/Ollama 等）')
    print('=' * 60)

    # ===== Step 1: 连接邮箱 =====
    try:
        conn = imaplib.IMAP4_SSL(imap_server, imap_port)
        conn.login(email_addr, auth_code)
        conn.select('INBOX')
    except Exception as e:
        print(f'❌ 邮箱连接失败: {e}')
        print(f'   服务器: {imap_server}:{imap_port}')
        print(f'   账号: {email_addr}')
        return

    # ===== Step 2: 日期过滤 + 获取邮件 =====
    print('\n📅 是否指定日期范围？（直接回车使用最近60封）')
    start = input('  开始日期 (YYYY-MM-DD): ').strip()
    end = input('  结束日期 (YYYY-MM-DD): ').strip()

    if start or end:
        parts = []
        display = []
        if start:
            try:
                dt = datetime.strptime(start, '%Y-%m-%d')
                parts.append(f'SINCE {dt.strftime("%d-%b-%Y")}')
                display.append(start)
            except ValueError:
                print('  ⚠️ 开始日期格式错误，已忽略')
        if end:
            try:
                dt = datetime.strptime(end, '%Y-%m-%d')
                parts.append(f'BEFORE {dt.strftime("%d-%b-%Y")}')
                display.append(end)
            except ValueError:
                print('  ⚠️ 结束日期格式错误，已忽略')
        search_cmd = ' '.join(parts) if parts else 'ALL'
        scan_label = '~'.join(display) if display else '全部'
    else:
        search_cmd = 'ALL'
        scan_label = '最近60封'

    status, data = conn.search(None, search_cmd)
    if status != 'OK' or not data[0]:
        print('收件箱为空')
        conn.logout()
        return

    mail_ids = data[0].split()
    if not start and not end:
        mail_ids = mail_ids[-60:]

    print(f'\n📬 扫描范围: {scan_label} ({len(mail_ids)} 封邮件)')

    # ===== Step 3: 分类 + 提取 =====
    records = []
    llm_count = 0
    rule_count = 0

    for idx, mid in enumerate(reversed(mail_ids), 1):
        status, msg_data = conn.fetch(mid, '(RFC822)')
        if status != 'OK':
            continue

        msg = email.message_from_bytes(msg_data[0][1])
        sender = decode_str(msg['From'])
        subject = decode_str(msg['Subject'])
        body = get_email_text(msg)
        attachments = save_attachments(msg, idx)

        # 分类（LLM 或规则）
        classify_result = classify_email(subject, sender, body[:500])

        category = classify_result['category']
        method = classify_result.get('method', '规则')
        if category == '不相关':
            continue

        # 提取（LLM 或规则）
        extract_result = extract_record(body, subject, category)

        # 机票按 PDF 拆分
        if category == '机票':
            pdfs = _get_pdf_attachments_raw(msg)
            if pdfs:
                for pdf in pdfs:
                    pdf_info = _parse_pdf_filename(pdf['filename'])
                    record = {**extract_result}
                    if pdf_info.get('金额'):
                        record['金额'] = pdf_info['金额']
                    if pdf_info.get('日期'):
                        record['日期'] = pdf_info['日期']
                    if pdf_info.get('出发地'):
                        record['出发地'] = pdf_info['出发地']
                    if pdf_info.get('目的地'):
                        record['目的地'] = pdf_info['目的地']
                    record['附件'] = '; '.join(attachments) if attachments else ''
                    record['主题'] = subject
                    record['发件人'] = sender
                    records.append(record)
                continue

        extract_result['附件'] = '; '.join(attachments) if attachments else ''
        extract_result['主题'] = subject
        extract_result['发件人'] = sender
        records.append(extract_result)

        # 统计
        if extract_result.get('方法', '').startswith('LLM'):
            llm_count += 1
        else:
            rule_count += 1

        short_subj = subject[:35] + '...' if len(subject) > 35 else subject
        amt = extract_result.get('金额', '') or '-'
        print(f'  [{idx:3d}] {category}  {short_subj}  ¥{amt}')

    conn.logout()

    if not records:
        print('\n未发现出差相关邮件')
        return

    # ===== Step 4: 出差聚合 =====
    trips = aggregate_trips(records)

    # ===== Step 5: 生成 Excel =====
    total_amount = sum(r.get('金额', 0) or 0 for r in records)
    _generate_excel(records, trips, total_amount, scan_label)

    # ===== 打印结果 =====
    print(f'\n{"=" * 60}')
    print(f'✅ 处理完成')
    print(f'  📧 出差相关: {len(records)} 条记录')
    if use_llm:
        print(f'  🧠 LLM 提取: {llm_count} 条  📋 规则降级: {rule_count} 条')
    print(f'  ✈️  识别到 {len(trips)} 次出差/旅行')
    print(f'  💰 总金额: ¥{total_amount:,.2f}')
    report_name = f"差旅汇总_{datetime.now().strftime('%Y%m%d')}.xlsx"
    print(f'  📊 Excel: {os.path.join(OUTPUT_DIR, report_name)}')
    print(f'  📎 附件: {ATTACH_DIR}/')

    if trips:
        print(f'\n📋 出差行程汇总:')
        for t in trips:
            print(f'  🏷️  Trip #{t["trip_id"]}  {t["summary"]}')
            print(f'     ¥{t["total"]:,.2f}  ({len(t["records"])} 条记录)')


def _get_pdf_attachments_raw(msg):
    """提取邮件中所有 PDF 附件"""
    pdfs = []
    if not msg.is_multipart():
        return pdfs
    for part in msg.walk():
        fn_raw = part.get_filename()
        if not fn_raw:
            continue
        fn = decode_str(fn_raw)
        ct = part.get_content_type()
        is_pdf = fn.lower().endswith('.pdf') or ct == 'application/pdf' or '.pdf' in fn.lower()
        if not is_pdf:
            continue
        pdfs.append({'filename': fn})
    return pdfs


def _parse_pdf_filename(fn):
    """从 PDF 文件名解析日期/路线/金额"""
    info = {'日期': '', '出发地': '', '目的地': '', '金额': ''}
    # 日期
    m = re.search(r'(\d{4})[-年](\d{1,2})[-月](\d{1,2})', fn)
    if m:
        info['日期'] = f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    # 路线
    m = re.search(r'([\u4e00-\u9fa5]{2,4})-([\u4e00-\u9fa5]{2,4})', fn)
    if m:
        info['出发地'] = m.group(1)
        info['目的地'] = m.group(2)
    # 金额
    m = re.search(r'(\d+\.\d{2})\.pdf$', fn)
    if m:
        info['金额'] = float(m.group(1))
    return info


def _generate_excel(records, trips, total_amount, scan_label):
    """生成 Excel 报表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    BLUE = '1A56DB'
    DARK_TEXT = '1F2937'

    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='微软雅黑')
    title_font = Font(bold=True, color=DARK_TEXT, size=14, name='微软雅黑')
    body_font = Font(size=10, name='微软雅黑')
    total_font = Font(bold=True, color=BLUE, size=11, name='微软雅黑')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    cat_fills = {
        '机票': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
        '酒店': PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid'),
        '网约车': PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        '发票': PatternFill(start_color='FCE7F3', end_color='FCE7F3', fill_type='solid'),
        '火车票': PatternFill(start_color='E0E7FF', end_color='E0E7FF', fill_type='solid'),
    }

    mode = 'AI增强模式' if llm_available() else '规则模式'
    total = sum(r.get('金额', 0) or 0 for r in records)

    # ===== Sheet 1: 报销总览 =====
    ws = wb.active
    ws.title = '报销总览'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'差旅费用报销汇总单 · {mode}'
    ws['A1'].font = title_font
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = f'生成日期：{datetime.now().strftime("%Y年%m月%d日")}  |  扫描范围：{scan_label}  |  出差相关：{len(records)} 条'
    ws['A2'].font = Font(size=9, color='6B7280', name='微软雅黑')

    # 总额卡片
    ws.merge_cells('A4:C4')
    ws['A4'] = '💰 可报销总额'
    ws['A4'].font = Font(size=12, color='6B7280', name='微软雅黑')
    ws.merge_cells('D4:F4')
    ws['D4'] = f'¥ {total:,.2f}'
    ws['D4'].font = Font(bold=True, size=22, color=BLUE, name='微软雅黑')
    ws['D4'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[4].height = 40

    # 出差行程汇总（新功能）
    if trips:
        row = 6
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = '📋 出差行程汇总'
        ws[f'A{row}'].font = Font(bold=True, size=11, color=DARK_TEXT, name='微软雅黑')

        trip_headers = ['行程编号', '目的地', '起止日期', '订单数', '金额合计', '摘要']
        trip_widths = [12, 12, 20, 10, 14, 30]
        row += 1
        for col, (h, w) in enumerate(zip(trip_headers, trip_widths), 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w

        for t in trips:
            row += 1
            vals = [
                f'Trip #{t["trip_id"]}',
                t['destination'],
                f'{t["start_date"]} ~ {t["end_date"]}',
                len(t['records']),
                t['total'],
                t['summary'],
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = body_font
                cell.alignment = center_align if col <= 4 else left_align
                cell.border = thin_border

    # 分类汇总
    cat_start = row + 2 if trips else 8
    ws.merge_cells(f'A{cat_start}:F{cat_start}')
    ws[f'A{cat_start}'] = '📊 按类别汇总'
    ws[f'A{cat_start}'].font = Font(bold=True, size=11, color=DARK_TEXT, name='微软雅黑')

    cat_start += 1
    cat_headers = ['类别', '笔数', '金额(元)', '占比', '', '']
    for col, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=cat_start, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    categories = {}
    for r in records:
        if r.get('金额', '') != '':
            cat = r.get('分类', '其他')
            categories[cat] = categories.get(cat, 0) + r['金额']

    row = cat_start
    for cat, amt in sorted(categories.items(), key=lambda x: -x[1]):
        row += 1
        count = sum(1 for r in records if r.get('分类') == cat and r.get('金额', '') != '')
        pct = f'{amt/total*100:.1f}%' if total else '0.0%'
        vals = [cat, count, amt, pct, '', '']
        cf = cat_fills.get(cat)
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.alignment = center_align
            cell.border = thin_border
            if cf:
                cell.fill = cf

    # 合计行
    row += 1
    total_count = sum(1 for r in records if r.get('金额', '') != '')
    for col, val in enumerate(['合计', total_count, total, '100%', '', ''], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = total_font
        cell.fill = PatternFill(start_color='E8EFFB', end_color='E8EFFB', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border

    # ===== Sheet 2: 费用明细 =====
    ws2 = wb.create_sheet('费用明细')
    ws2.merge_cells('A1:H1')
    ws2['A1'] = '差旅费用明细表'
    ws2['A1'].font = title_font
    ws2.row_dimensions[1].height = 28

    detail_headers = ['序号', '日期', '类别', '供应商/平台', '金额(元)', '出发地→目的地', '方法', '附件']
    detail_widths = [6, 14, 10, 18, 12, 18, 12, 20]
    for col, (h, w) in enumerate(zip(detail_headers, detail_widths), 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col)].width = w

    sorted_records = sorted(records, key=lambda r: str(r.get('日期', '')), reverse=True)

    for i, r in enumerate(sorted_records, 1):
        row = i + 2
        route = f"{r.get('出发地', '')}→{r.get('目的地', '')}" if r.get('出发地') and r.get('目的地') else ''
        vals = [
            i,
            r.get('日期', '') or '-',
            r.get('分类', ''),
            r.get('酒店名称', '') or r.get('供应商', '') or r.get('平台', '') or '-',
            r.get('金额', '') if r.get('金额', '') != '' else '-',
            route or '-',
            r.get('方法', '') or '-',
            r.get('附件', '') or '-',
        ]
        cf = cat_fills.get(r.get('分类', ''))
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.alignment = center_align if col in (1, 3, 5, 7) else left_align
            cell.border = thin_border
            if col == 3 and cf:
                cell.fill = cf

    # ===== Sheet 3: 按供应商 =====
    ws3 = wb.create_sheet('按供应商')
    ws3.merge_cells('A1:D1')
    ws3['A1'] = f'按供应商/平台统计 · {mode}'
    ws3['A1'].font = title_font
    ws3.row_dimensions[1].height = 28

    vendor_headers = ['供应商', '笔数', '金额(元)', '占比']
    vendor_widths = [20, 10, 14, 10]
    for col, (h, w) in enumerate(zip(vendor_headers, vendor_widths), 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws3.column_dimensions[get_column_letter(col)].width = w

    vendors = {}
    for r in records:
        v = r.get('酒店名称', '') or r.get('供应商', '') or r.get('平台', '') or '其他'
        amt = r.get('金额', 0) or 0
        if v not in vendors:
            vendors[v] = {'count': 0, 'amount': 0}
        vendors[v]['count'] += 1
        vendors[v]['amount'] += amt

    for i, (v, d) in enumerate(sorted(vendors.items(), key=lambda x: -x[1]['amount']), 1):
        row = i + 2
        vals = [v, d['count'], d['amount'], f"{d['amount']/max(total, 1)*100:.1f}%"]
        rf = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid') if i % 2 == 0 else PatternFill()
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.font = body_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.fill = rf

    today = datetime.now().strftime('%Y%m%d')
    xlsx_path = os.path.join(OUTPUT_DIR, f'差旅汇总_{today}.xlsx')
    wb.save(xlsx_path)


if __name__ == '__main__':
    main()
